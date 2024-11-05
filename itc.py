import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import time
import threading
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

############################################################################################
########### Directorios
############################################################################################


# Crear la ventana principal
root = tk.Tk()
root.title("ITC Automatización")
root.geometry("700x300")

# Etiqueta para el título
label_title = tk.Label(root, text="Selecciona un directorio principal", font=("Arial", 16))
label_title.pack(pady=20)

# Etiqueta para mostrar el directorio seleccionado
label_dir = tk.Label(root, text="", font=("Arial", 10))
label_dir.pack(pady=10)

# Variables globales para rutas
selected_directory = None
csv_file_path = None
excel_file_path = None

# Función para seleccionar el directorio
def select_directory():
    global selected_directory, csv_file_path, excel_file_path
    selected_directory = filedialog.askdirectory()
    
    if selected_directory:
        # Mostrar el directorio principal seleccionado
        label_dir.config(text=f"Directorio principal seleccionado: {selected_directory}")
        
        # Buscar archivo CSV en el directorio principal
        for file in os.listdir(selected_directory):
            if file.endswith(".csv"):
                csv_file_path = os.path.join(selected_directory, file)
                break
        else:
            messagebox.showerror("Error", "No se encontró ningún archivo CSV en el directorio seleccionado")
            return
        
        # Buscar archivo Excel en la subcarpeta 'data'
        excel_file_path = os.path.join(selected_directory, "data", "itc.xlsx")
        if not os.path.isfile(excel_file_path):
            messagebox.showerror("Error", "No se encontró el archivo itc.xlsx en la carpeta 'data'")
            return
        
        # Mostrar el botón de Aceptar para cerrar la ventana
        btn_accept.pack(pady=10)

# Función para cerrar la ventana
def close_popup():
    root.destroy()

# Botón para seleccionar el directorio
btn_select_dir = tk.Button(root, text="Seleccionar Directorio", command=select_directory)
btn_select_dir.pack(pady=10)

# Botón de Aceptar que se mostrará después de seleccionar el directorio
btn_accept = tk.Button(root, text="Aceptar", command=close_popup)

# Ejecutar la ventana principal
root.mainloop()

############################################################################################
########### Logica
############################################################################################

def rename_unnamed_columns(df):
    col_counter = 1

    new_columns = []
    for col in df.columns:
        if 'Unnamed' in col:
            new_columns.append(f'{col_counter}')
            col_counter += 1
        else:
            new_columns.append(col)

    # Asignamos los nuevos nombres de columna al DataFrame
    df.columns = new_columns
    return df

df = pd.read_csv(csv_file_path, delimiter=';')

df = rename_unnamed_columns(df)

############################################################################################
########### Tabla 1

work_999_index = df[df.iloc[:, 0].str.contains('Work.999', na=False)].index

if not work_999_index.empty:
    next_row_index = work_999_index[0] + 1
    tabla1 = df.iloc[next_row_index]

Nombre_Inter = tabla1.iloc[0]
IP =  tabla1.iloc[2]
ID =  tabla1.iloc[3]
Num_grupos =  tabla1.iloc[6]
Num_E_Logicos =  tabla1.iloc[8]
Num_S_Logicos =  tabla1.iloc[7]
Num_paneles =  tabla1.iloc[4]
Num_D_Logicos =  tabla1.iloc[12]

data = {
    'Dato': [
        'Nombre de la intersección',
        'IP controlador',
        'ID panel de control',
        'Número de grupos',
        'Cantidad de escenarios lógicos',
        'Cantidad de secuencias lógicas',
        'Número de planes',
        'Número de detectores lógicos'
    ],
    'Valor': [
        Nombre_Inter,
        IP,
        ID,
        Num_grupos,
        Num_E_Logicos,
        Num_S_Logicos,
        Num_paneles,
        Num_D_Logicos
    ]
}

df1 = pd.DataFrame(data)

################################### Completar datos
sheet_name = "Tabla 1-1"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

ws["B2"] = df1.iloc[0,1]
ws["B3"] = df1.iloc[1,1]
ws["B4"] = df1.iloc[2,1]
ws["B5"] = df1.iloc[3,1]
ws["B6"] = df1.iloc[4,1]
ws["B7"] = df1.iloc[5,1]
ws["B8"] = df1.iloc[6,1]
ws["B9"] = df1.iloc[7,1]

wb.save(excel_file_path)



############################################################################################
########### Tabla 2

work_998_index = df[df.iloc[:, 0].str.contains('Work.998', na=False)].index

if not work_998_index.empty and not work_999_index.empty:
    tabla2 = df.iloc[work_998_index[0] +1 : work_999_index[0]-1] #quito el work y el Next
else:
    print("No se encontraron ambas entradas de 'Work.998' y 'Work.999'")

df2 = pd.DataFrame({
    'Grupo': pd.Series(dtype='str'),   # Columna vacía 'Grupo'
    'Nombre': pd.Series(dtype='str'),  # Columna vacía 'Nombre'
    'Tiempos1': pd.Series(dtype='object'),  # Columna vacía 'Tiempos1'
    'Tiempos2': pd.Series(dtype='object'),  # Columna vacía 'Tiempos2'
    'Tiempos3': pd.Series(dtype='object')   # Columna vacía 'Tiempos3'
})


resultados = []
num=1

for idx, row in tabla2.iterrows():
    valores = [row[0], row[11], row[14], row[9]]

    # Insertar el número de fila al inicio de la lista
    fila_data = [num] + valores  # +1 porque queremos que la primera fila sea 1
    num = num +1
    resultados.append(fila_data)

# Convertir la lista de resultados en un DataFrame
df_nuevo = pd.DataFrame(resultados, columns=['Fila', 'Dato1', 'Dato2', 'Dato3', 'Dato4'])

df2['Grupo'] = df_nuevo['Fila'].values.tolist()
df2['Nombre'] = df_nuevo[ 'Dato1'].values.tolist()
df2['Tiempos1'] = df_nuevo['Dato2'].values.tolist()
df2['Tiempos2'] = df_nuevo['Dato3'].values.tolist()
df2['Tiempos3'] = df_nuevo['Dato4'].values.tolist()


def procesar_valor(valor):
    # Realizamos el split y tomamos el primer elemento
    primer_valor = valor.split('-')[0]
    # Convertimos a entero
    return int(float(primer_valor))

# Aplicar la función para todas las columnas de tiempos
df2['Tiempos1'] = df2['Tiempos1'].apply(procesar_valor)
df2['Tiempos2'] = df2['Tiempos2'].apply(procesar_valor)
df2['Tiempos3'] = df2['Tiempos3'].apply(procesar_valor)


################################### Completar datos
sheet_name = "Tabla 1-2"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

# GRUPO
for i in range(len(df2)):
    ws[f"A{i + 2}"] = df2.iloc[i, 0]  # Escribir en la columna B, fila dinámica (i + 1)

# NOMBRE
for i in range(len(df2)):
    ws[f"B{i + 2}"] = df2.iloc[i, 1]  

# TIEMPOS 1
for i in range(len(df2)):
    ws[f"C{i + 2}"] = df2.iloc[i, 2]  

# TIEMPOS 2
for i in range(len(df2)):
    ws[f"D{i + 2}"] = df2.iloc[i, 3]  

# TIEMPOS 3
for i in range(len(df2)):
    ws[f"E{i + 2}"] = df2.iloc[i, 4]  

wb.save(excel_file_path)



############################################################################################
########### Tabla 3

work_997_index = df[df.iloc[:, 0].str.contains('Work.997', na=False)].index

if not work_997_index.empty and not work_998_index.empty:
    tabla3 = df.iloc[work_997_index[0] +1 : work_998_index[0]-1] #quito el work y el Next
else:
    print("No se encontraron ambas entradas de 'Work.997' y 'Work.998'")

grupo_vals = df2['Nombre'].values
# Crear una tabla simétrica vacía usando estos valores
tabla_simetrica = pd.DataFrame(index=grupo_vals, columns=grupo_vals)

for i in range(len(tabla_simetrica.columns)):
    tabla_simetrica.iloc[:, i] = tabla3.iloc[:, i].values

def convertir_a_int(valor):
    try:
        return int(float(valor))  # Convertir a float y luego a int
    except (ValueError, TypeError):
        return valor  # Si no se puede convertir, devolver el valor original

tabla_simetrica = tabla_simetrica.applymap(convertir_a_int)

def eliminar_espacios(valor):
    if isinstance(valor, str):  # Verificar si el valor es una cadena de texto
        return valor.strip()  # Eliminar espacios en blanco al inicio y al final
    return valor  # Si no es una cadena, devolver el valor tal como está

# Aplicar la función a todo el DataFrame
tabla_simetrica = tabla_simetrica.applymap(eliminar_espacios)
tabla_simetrica = tabla_simetrica.replace('.', '-')
tabla_simetrica = tabla_simetrica.fillna('-')

################################### Completar datos
sheet_name = "Tabla 1-3"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

# GRUPO VERTICAL
for i in range(len(df2)):
    ws[f"B{i + 3}"] = df2.iloc[i,1]

# GRUPO HORIZONTAL
for i in range(len(df2)):
    col_letra = get_column_letter(3 + i)  # Empieza en la columna C (columna 3)
    ws[f"{col_letra}2"] = df2.iloc[i, 1]  # Escribir en la columna dinámica y fila fija 3

# Iterar sobre todas las filas y columnas de tabla_simetrica
for col_idx in range(len(tabla_simetrica.columns)):
    col_letter = get_column_letter(3 + col_idx)  # 3 corresponde a la columna "C" y avanzamos
    for row_idx in range(len(tabla_simetrica)):
        ws[f"{col_letter}{row_idx + 3}"] = tabla_simetrica.iloc[row_idx, col_idx]


wb.save(excel_file_path)
############################################################################################
########### Tabla 4

work_006_index = df[df.iloc[:, 0].str.contains('Work.006', na=False)].index
work_007_index = df[df.iloc[:, 0].str.contains('Work.007', na=False)].index

if not work_006_index.empty and not work_007_index.empty:
    tabla4 = df.iloc[work_006_index[0] +1 : work_007_index[0]-1] #quito el work y el Next
else:
    print("No se encontraron ambas entradas de 'Work.997' y 'Work.998'")

grupo_vals = df2['Nombre'].values
# Crear una tabla simétrica vacía usando estos valores
tabla_simetrica2 = pd.DataFrame(index=grupo_vals, columns=grupo_vals)

for i in range(len(tabla_simetrica.columns)):
    tabla_simetrica2.iloc[:, i] = tabla4.iloc[:, i].values

def convertir_a_int(valor):
    try:
        primer_valor = valor.split('-')[1]
        return int(float(primer_valor))  # Convertir a float y luego a int
    except (ValueError, TypeError):
        return valor  # Si no se puede convertir, devolver el valor original

tabla_simetrica2 = tabla_simetrica2.applymap(convertir_a_int)

def eliminar_puntos(valor):
    if isinstance(valor, str):  # Verificar si el valor es una cadena de texto
        return valor.replace(".", "")  # Eliminar todos los espacios en blanco
    return valor  # Si no es una cadena, devolver el valor tal como está

# Aplicar la función a todo el DataFrame
df4 = tabla_simetrica2.applymap(eliminar_puntos)


################################### Completar datos
sheet_name = "Tabla 1-4"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

# GRUPO VERTICAL
for i in range(len(df2)):
    ws[f"B{i + 3}"] = df2.iloc[i,1]

# GRUPO HORIZONTAL
for i in range(len(df2)):
    col_letra = get_column_letter(3 + i)  # Empieza en la columna C (columna 3)
    ws[f"{col_letra}2"] = df2.iloc[i, 1]  # Escribir en la columna dinámica y fila fija 3

# Iterar sobre todas las filas y columnas de tabla_simetrica
for col_idx in range(len(df4.columns)):
    col_letter = get_column_letter(3 + col_idx)  # 3 corresponde a la columna "C" y avanzamos
    for row_idx in range(len(df4)):
        ws[f"{col_letter}{row_idx + 3}"] = df4.iloc[row_idx, col_idx]

wb.save(excel_file_path)
############################################################################################
########### Tabla 5

work_009_index = df[df.iloc[:, 0].str.contains('Work.009', na=False)].index
work_010_index = df[df.iloc[:, 0].str.contains('Work.010', na=False)].index

if not work_009_index.empty and not work_010_index.empty:
    tabla5 = df.iloc[work_009_index[0] +1 : work_010_index[0]-1] #quito el work y el Next
else:
    print("No se encontraron ambas entradas de 'Work.997' y 'Work.998'")

segunda_columna_lista = tabla5.iloc[:, 1].tolist()


fase_vals = [1, 2, 3, 4]
grupo_vals = df2['Nombre'].values

df5 = pd.DataFrame({'fase': fase_vals})


for grupo in grupo_vals:
    df5[grupo] = pd.Series([None] * len(fase_vals))  # Inicializamos las columnas con valores vacíos

segunda_columna_lista = list(map(int, segunda_columna_lista))

# Definir las potencias de 2 que vamos a usar
potencias = [16, 8, 4, 2]

# Función para descomponer el valor en potencias de 2 más grandes posibles
def descomponer_en_potencias(valor):
    resultado = []
    for potencia in potencias:
        if valor >= potencia:
            resultado.append(potencia)
            valor -= potencia  # Restamos la potencia del valor
    return resultado

# Función para llenar el DataFrame
def llenar_columna(valor, nombre_columna):
    # Descomponemos el valor en las potencias de 2 correspondientes
    potencias_descompuestas = descomponer_en_potencias(valor)

    # Llenamos el DataFrame según las potencias descompuestas
    for potencia in potencias_descompuestas:
        # Fila 1 corresponde a 2, fila 2 a 4, fila 3 a 8, fila 4 a 16
        if potencia == 2:
            df5.loc[0, nombre_columna] = 'ok'  # Fila 1
        elif potencia == 4:
            df5.loc[1, nombre_columna] = 'ok'  # Fila 2
        elif potencia == 8:
            df5.loc[2, nombre_columna] = 'ok'  # Fila 3
        elif potencia == 16:
            df5.loc[3, nombre_columna] = 'ok'  # Fila 4

# Recorrer la lista de valores binarios y llenar las columnas
for idx, valor in enumerate(segunda_columna_lista):
    nombre_columna = grupo_vals[idx]  # Usamos el nombre de la columna directamente
    llenar_columna(int(valor), nombre_columna) 

df5.dropna(how='all', inplace=True)

################################### Completar datos
sheet_name = "Tabla 1-5"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]


# GRUPO HORIZONTAL
for i in range(len(df2)):
    col_letra = get_column_letter(2 + i)  
    ws[f"{col_letra}2"] = df2.iloc[i, 1]  

# Iterar sobre todas las filas y columnas de tabla_simetrica
# Obtener el valor y el estilo de la celda J14
valor_w1 = ws["W1"].value
estilo_w1 = ws["W1"]._style  # El estilo de la celda J14

# Recorrer las columnas y filas de df5
for col_idx in range(len(df5.columns)):
    col_letter = get_column_letter(1 + col_idx)  # Comienza en la columna "A"
    for row_idx in range(len(df5)):
        if df5.iloc[row_idx, col_idx] == "ok":
            # Obtener la dirección de la celda de destino
            destino = f"{col_letter}{row_idx + 3}"

            # Asignar el valor de J14 a la celda de destino
            ws[destino].value = valor_w1

            # Copiar el formato desde J14
            ws[destino]._style = estilo_w1  # Aplicar el estilo de J14 a la celda de destino


wb.save(excel_file_path)
############################################################################################
########### Tabla 6

######################### TABLA 1 (CONTIENE TODAS LAS SUB TABLAS)
work_008_index = df[df.iloc[:, 0].str.contains('Work.008', na=False)].index

if not work_008_index.empty and not work_009_index.empty:
    tabla6 = df.iloc[work_008_index[0] +1 : work_009_index[0]] #quito el work y el Next
else:
    print("No se encontraron ambas entradas de 'Work.997' y 'Work.998'")

######################### Contamos cuantas sub tablas hay
secuencias = (tabla6.iloc[:, 0] == "NeXt").sum()
secuencias

######################### TABLA 2
if secuencias >= 2:
    index_next = tabla6[tabla6.iloc[:, 0] == "NeXt"].index[0]
    tabla62 = tabla6.loc[index_next + 1:].reset_index(drop=True)

######################### TABLA 3
if secuencias >= 3:
  index_next = tabla62[tabla62.iloc[:, 0] == "NeXt"].index[0]
  tabla63 = tabla62.loc[index_next + 1:].reset_index(drop=True)

######################### TABLA 4
if secuencias >= 4:
  index_next = tabla63[tabla63.iloc[:, 0] == "NeXt"].index[0]
  tabla64 = tabla63.loc[index_next + 1:].reset_index(drop=True)

######################### TABLA 5
if secuencias >= 5:
  index_next = tabla64[tabla64.iloc[:, 0] == "NeXt"].index[0]
  tabla65 = tabla64.loc[index_next + 1:].reset_index(drop=True)

################################################################################
################## Creamos los df final
################################################################################
fase_vals = ["Permanente", "Por detectores", "Con fase o por detectores", "Verde máximo mínimo permanente", "Verde máximo mínimo por detector", "Verde máximo permanente", "Verde máximo por detector","Cuenta regresiva comienza cuando un grupo en conflicto es demandado", "Cuenta regresiva comienza sin que un grupo en conflicto sea demandado","Luego de la extensión máxima de verde","Luego de la extensión máxima de verde y verde pasivo"]

df6 = pd.DataFrame()
df62 = pd.DataFrame()
df63 = pd.DataFrame()
df64 = pd.DataFrame()
df65 = pd.DataFrame()


################################################################
################## Primer DF
################################################################
if secuencias >= 1:
  # Inicializamos las columnas en el DataFrame
  for grupo in grupo_vals:
      df6[grupo] = pd.Series([None] * len(fase_vals))  # Columnas vacías

  df6.index = fase_vals

  ########################### Completamos la primera parte
  lista_valores1_1 = []

  for valor in tabla6.iloc[:, 0]:  # Recorre la primera columna
      if valor == "NeXt":
          break
      lista_valores1_1.append(valor)  # Agrega los valores a la list

  for i in range(len(lista_valores1_1)):
    if pd.isna(lista_valores1_1[i]):  # Verificar si es NaN
        lista_valores1_1[i] = "Por detectores"
    elif lista_valores1_1[i] == "1":
        lista_valores1_1[i] = "Permanente"
    elif lista_valores1_1[i] == "2":
        lista_valores1_1[i] = "Con fase o por detectores"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
    fila_a_modificar = lista_valores1_1[i]  # Obtener el valor de la lista en la posición i
    df6.loc[fila_a_modificar, columna] = 'ok'


  ########################### Completamos la segunda parte
  lista_valores2_1 = []

  for valor_col_0, valor_col_4 in zip(tabla6.iloc[:, 0], tabla6.iloc[:, 4]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores2_1.append(valor_col_4)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores2_1)):
      if pd.isna(lista_valores2_1[i]):  # Verificar si es NaN
          lista_valores2_1[i] = "Verde máximo mínimo por detector"
      elif lista_valores2_1[i] == "1":
          lista_valores2_1[i] = "Verde máximo mínimo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores2_1[i]  # Obtener el valor de la lista en la posición i
      df6.loc[fila_a_modificar, columna] = 'ok'


  ########################### Completamos la tercera parte
  lista_valores3_1 = []

  for valor_col_0, valor_col_5 in zip(tabla6.iloc[:, 0], tabla6.iloc[:, 5]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores3_1.append(valor_col_5)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores3_1)):
      if pd.isna(lista_valores3_1[i]):  # Verificar si es NaN
          lista_valores3_1[i] = "Verde máximo por detector"
      elif lista_valores3_1[i] == "1":
          lista_valores3_1[i] = "Verde máximo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores3_1[i]  # Obtener el valor de la lista en la posición i
      df6.loc[fila_a_modificar, columna] = 'ok'


  ########################### Completamos la cuarta parte
  lista_valores4_1 = []

  for valor_col_0, valor_col_7 in zip(tabla6.iloc[:, 0], tabla6.iloc[:, 7]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores4_1.append(valor_col_7)  # Agregamos a la lista los valores de la columna 4

  # Recorremos la lista y alteramos solo los campos de texto
  for i in range(len(lista_valores4_1)):
      if isinstance(lista_valores4_1[i], str):  # Verificamos si es una cadena de texto
          lista_valores4_1[i] = lista_valores4_1[i].split('-')[0]

  for i in range(len(lista_valores4_1)):
      if pd.isna(lista_valores4_1[i]):  # Verificar si es NaN
          lista_valores4_1[i] = "Cuenta regresiva comienza cuando un grupo en conflicto es demandado"
      elif lista_valores4_1[i] == "1":
          lista_valores4_1[i] = "Cuenta regresiva comienza sin que un grupo en conflicto sea demandado"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores4_1[i]  # Obtener el valor de la lista en la posición i
      df6.loc[fila_a_modificar, columna] = 'ok'


  ########################### Completamos la quinta parte
  lista_valores5_1 = []

  for valor_col_0, valor_col_2 in zip(tabla6.iloc[:, 0], tabla6.iloc[:, 2]):
      if valor_col_0 == "NeXt":
          break
      lista_valores5_1.append(valor_col_2)

  for i in range(len(lista_valores5_1)):
      if pd.isna(lista_valores5_1[i]):  # Verificar si es NaN
          lista_valores5_1[i] = "Luego de la extensión máxima de verde y verde pasivo"
      elif lista_valores5_1[i] == "3":
          lista_valores5_1[i] = "Luego de la extensión máxima de verde"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores5_1[i]  # Obtener el valor de la lista en la posición i
      df6.loc[fila_a_modificar, columna] = 'ok'

  df6 = df6.fillna("-")
  df61 = df6

################################################################
################## Segundo DF
################################################################
if secuencias >= 2:
  # Inicializamos las columnas en el DataFrame
  for grupo in grupo_vals:
      df62[grupo] = pd.Series([None] * len(fase_vals))  # Columnas vacías

  df62.index = fase_vals

  ########################### Completamos la primera parte
  lista_valores1_2 = []

  for valor in tabla62.iloc[:, 0]:  # Recorre la primera columna
      if valor == "NeXt":
          break
      lista_valores1_2.append(valor)  # Agrega los valores a la list

  for i in range(len(lista_valores1_2)):
    if pd.isna(lista_valores1_2[i]):  # Verificar si es NaN
        lista_valores1_2[i] = "Por detectores"
    elif lista_valores1_2[i] == "1":
        lista_valores1_2[i] = "Permanente"
    elif lista_valores1_2[i] == "2":
        lista_valores1_2[i] = "Con fase o por detectores"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
    fila_a_modificar = lista_valores1_2[i]  # Obtener el valor de la lista en la posición i
    df62.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la segunda parte
  lista_valores2_2 = []

  for valor_col_0, valor_col_4 in zip(tabla62.iloc[:, 0], tabla62.iloc[:, 4]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores2_2.append(valor_col_4)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores2_2)):
      if pd.isna(lista_valores2_2[i]):  # Verificar si es NaN
          lista_valores2_2[i] = "Verde máximo mínimo por detector"
      elif lista_valores2_2[i] == "1":
          lista_valores2_2[i] = "Verde máximo mínimo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores2_2[i]  # Obtener el valor de la lista en la posición i
      df62.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la tercera parte
  lista_valores3_2 = []

  for valor_col_0, valor_col_5 in zip(tabla62.iloc[:, 0], tabla62.iloc[:, 5]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores3_2.append(valor_col_5)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores3_2)):
      if pd.isna(lista_valores3_2[i]):  # Verificar si es NaN
          lista_valores3_2[i] = "Verde máximo por detector"
      elif lista_valores3_2[i] == "1":
          lista_valores3_2[i] = "Verde máximo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores3_2[i]  # Obtener el valor de la lista en la posición i
      df62.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la cuarta parte
  lista_valores4_2 = []

  for valor_col_0, valor_col_7 in zip(tabla62.iloc[:, 0], tabla62.iloc[:, 7]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores4_2.append(valor_col_7)  # Agregamos a la lista los valores de la columna 4

  # Recorremos la lista y alteramos solo los campos de texto
  for i in range(len(lista_valores4_2)):
      if isinstance(lista_valores4_2[i], str):  # Verificamos si es una cadena de texto
          lista_valores4_2[i] = lista_valores4_2[i].split('-')[0]

  for i in range(len(lista_valores4_2)):
      if pd.isna(lista_valores4_2[i]):  # Verificar si es NaN
          lista_valores4_2[i] = "Cuenta regresiva comienza cuando un grupo en conflicto es demandado"
      elif lista_valores4_2[i] == "1":
          lista_valores4_2[i] = "Cuenta regresiva comienza sin que un grupo en conflicto sea demandado"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores4_2[i]  # Obtener el valor de la lista en la posición i
      df62.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la quinta parte
  lista_valores5_2 = []

  for valor_col_0, valor_col_2 in zip(tabla62.iloc[:, 0], tabla62.iloc[:, 2]):
      if valor_col_0 == "NeXt":
          break
      lista_valores5_2.append(valor_col_2)

  for i in range(len(lista_valores5_2)):
      if pd.isna(lista_valores5_2[i]):  # Verificar si es NaN
          lista_valores5_2[i] = "Luego de la extensión máxima de verde y verde pasivo"
      elif lista_valores5_2[i] == "3":
          lista_valores5_2[i] = "Luego de la extensión máxima de verde"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores5_2[i]  # Obtener el valor de la lista en la posición i
      df62.loc[fila_a_modificar, columna] = 'ok'

  df62 = df62.fillna("-")

################################################################
################## Tercer DF
################################################################
if secuencias >= 3:
  # Inicializamos las columnas en el DataFrame
  for grupo in grupo_vals:
      df63[grupo] = pd.Series([None] * len(fase_vals))  # Columnas vacías

  df63.index = fase_vals

  ########################### Completamos la primera parte
  lista_valores1_3 = []

  for valor in tabla63.iloc[:, 0]:  # Recorre la primera columna
      if valor == "NeXt":
          break
      lista_valores1_3.append(valor)  # Agrega los valores a la list

  for i in range(len(lista_valores1_3)):
    if pd.isna(lista_valores1_3[i]):  # Verificar si es NaN
        lista_valores1_3[i] = "Por detectores"
    elif lista_valores1_3[i] == "1":
        lista_valores1_3[i] = "Permanente"
    elif lista_valores1_3[i] == "2":
        lista_valores1_3[i] = "Con fase o por detectores"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
    fila_a_modificar = lista_valores1_3[i]  # Obtener el valor de la lista en la posición i
    df63.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la segunda parte
  lista_valores2_3 = []

  for valor_col_0, valor_col_4 in zip(tabla63.iloc[:, 0], tabla63.iloc[:, 4]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores2_3.append(valor_col_4)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores2_3)):
      if pd.isna(lista_valores2_3[i]):  # Verificar si es NaN
          lista_valores2_3[i] = "Verde máximo mínimo por detector"
      elif lista_valores2_3[i] == "1":
          lista_valores2_3[i] = "Verde máximo mínimo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores2_3[i]  # Obtener el valor de la lista en la posición i
      df63.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la tercera parte
  lista_valores3_3 = []

  for valor_col_0, valor_col_5 in zip(tabla63.iloc[:, 0], tabla63.iloc[:, 5]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores3_3.append(valor_col_5)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores3_3)):
      if pd.isna(lista_valores3_3[i]):  # Verificar si es NaN
          lista_valores3_3[i] = "Verde máximo por detector"
      elif lista_valores3_3[i] == "1":
          lista_valores3_3[i] = "Verde máximo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores3_3[i]  # Obtener el valor de la lista en la posición i
      df63.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la cuarta parte
  lista_valores4_3 = []

  for valor_col_0, valor_col_7 in zip(tabla63.iloc[:, 0], tabla63.iloc[:, 7]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores4_3.append(valor_col_7)  # Agregamos a la lista los valores de la columna 4

  # Recorremos la lista y alteramos solo los campos de texto
  for i in range(len(lista_valores4_3)):
      if isinstance(lista_valores4_3[i], str):  # Verificamos si es una cadena de texto
          lista_valores4_3[i] = lista_valores4_3[i].split('-')[0]

  for i in range(len(lista_valores4_3)):
      if pd.isna(lista_valores4_3[i]):  # Verificar si es NaN
          lista_valores4_3[i] = "Cuenta regresiva comienza cuando un grupo en conflicto es demandado"
      elif lista_valores4_3[i] == "1":
          lista_valores4_3[i] = "Cuenta regresiva comienza sin que un grupo en conflicto sea demandado"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores4_3[i]  # Obtener el valor de la lista en la posición i
      df63.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la quinta parte
  lista_valores5_3 = []

  for valor_col_0, valor_col_2 in zip(tabla63.iloc[:, 0], tabla63.iloc[:, 2]):
      if valor_col_0 == "NeXt":
          break
      lista_valores5_3.append(valor_col_2)

  for i in range(len(lista_valores5_3)):
      if pd.isna(lista_valores5_3[i]):  # Verificar si es NaN
          lista_valores5_3[i] = "Luego de la extensión máxima de verde y verde pasivo"
      elif lista_valores5_3[i] == "3":
          lista_valores5_3[i] = "Luego de la extensión máxima de verde"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores5_3[i]  # Obtener el valor de la lista en la posición i
      df63.loc[fila_a_modificar, columna] = 'ok'

  df63 = df63.fillna("-")

################################################################
################## Cuarto DF
################################################################
if secuencias >= 4:
  # Inicializamos las columnas en el DataFrame
  for grupo in grupo_vals:
      df64[grupo] = pd.Series([None] * len(fase_vals))  # Columnas vacías

  df64.index = fase_vals

  ########################### Completamos la primera parte
  lista_valores1_4 = []

  for valor in tabla64.iloc[:, 0]:  # Recorre la primera columna
      if valor == "NeXt":
          break
      lista_valores1_4.append(valor)  # Agrega los valores a la list

  for i in range(len(lista_valores1_4)):
    if pd.isna(lista_valores1_4[i]):  # Verificar si es NaN
        lista_valores1_4[i] = "Por detectores"
    elif lista_valores1_4[i] == "1":
        lista_valores1_4[i] = "Permanente"
    elif lista_valores1_4[i] == "2":
        lista_valores1_4[i] = "Con fase o por detectores"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
    fila_a_modificar = lista_valores1_4[i]  # Obtener el valor de la lista en la posición i
    df64.loc[fila_a_modificar, columna] = 'ok'


  ########################### Completamos la segunda parte
  lista_valores2_4 = []

  for valor_col_0, valor_col_4 in zip(tabla64.iloc[:, 0], tabla64.iloc[:, 4]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores2_4.append(valor_col_4)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores2_4)):
      if pd.isna(lista_valores2_4[i]):  # Verificar si es NaN
          lista_valores2_4[i] = "Verde máximo mínimo por detector"
      elif lista_valores2_4[i] == "1":
          lista_valores2_4[i] = "Verde máximo mínimo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores2_4[i]  # Obtener el valor de la lista en la posición i
      df64.loc[fila_a_modificar, columna] = 'ok'


  ########################### Completamos la tercera parte
  lista_valores3_4 = []

  for valor_col_0, valor_col_5 in zip(tabla64.iloc[:, 0], tabla64.iloc[:, 5]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores3_4.append(valor_col_5)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores3_4)):
      if pd.isna(lista_valores3_3[i]):  # Verificar si es NaN
          lista_valores3_4[i] = "Verde máximo por detector"
      elif lista_valores3_4[i] == "1":
          lista_valores3_4[i] = "Verde máximo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores3_4[i]  # Obtener el valor de la lista en la posición i
      df64.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la cuarta parte
  lista_valores4_4 = []

  for valor_col_0, valor_col_7 in zip(tabla64.iloc[:, 0], tabla64.iloc[:, 7]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores4_4.append(valor_col_7)  # Agregamos a la lista los valores de la columna 4

  # Recorremos la lista y alteramos solo los campos de texto
  for i in range(len(lista_valores4_4)):
      if isinstance(lista_valores4_4[i], str):  # Verificamos si es una cadena de texto
          lista_valores4_4[i] = lista_valores4_4[i].split('-')[0]

  for i in range(len(lista_valores4_4)):
      if pd.isna(lista_valores4_4[i]):  # Verificar si es NaN
          lista_valores4_4[i] = "Cuenta regresiva comienza cuando un grupo en conflicto es demandado"
      elif lista_valores4_4[i] == "1":
          lista_valores4_4[i] = "Cuenta regresiva comienza sin que un grupo en conflicto sea demandado"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores4_4[i]  # Obtener el valor de la lista en la posición i
      df64.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la quinta parte
  lista_valores5_4 = []

  for valor_col_0, valor_col_2 in zip(tabla64.iloc[:, 0], tabla64.iloc[:, 2]):
      if valor_col_0 == "NeXt":
          break
      lista_valores5_4.append(valor_col_2)

  for i in range(len(lista_valores5_4)):
      if pd.isna(lista_valores5_4[i]):  # Verificar si es NaN
          lista_valores5_4[i] = "Luego de la extensión máxima de verde y verde pasivo"
      elif lista_valores5_4[i] == "3":
          lista_valores5_4[i] = "Luego de la extensión máxima de verde"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores5_4[i]  # Obtener el valor de la lista en la posición i
      df64.loc[fila_a_modificar, columna] = 'ok'

  df64 = df64.fillna("-")

################################################################
################## Quinto DF
################################################################
if secuencias >= 5:
  # Inicializamos las columnas en el DataFrame
  for grupo in grupo_vals:
      df65[grupo] = pd.Series([None] * len(fase_vals))  # Columnas vacías

  df65.index = fase_vals

  ########################### Completamos la primera parte
  lista_valores1_5 = []

  for valor in tabla65.iloc[:, 0]:  # Recorre la primera columna
      if valor == "NeXt":
          break
      lista_valores1_5.append(valor)  # Agrega los valores a la list

  for i in range(len(lista_valores1_5)):
    if pd.isna(lista_valores1_5[i]):  # Verificar si es NaN
        lista_valores1_5[i] = "Por detectores"
    elif lista_valores1_5[i] == "1":
        lista_valores1_5[i] = "Permanente"
    elif lista_valores1_5[i] == "2":
        lista_valores1_5[i] = "Con fase o por detectores"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
    fila_a_modificar = lista_valores1_5[i]  # Obtener el valor de la lista en la posición i
    df65.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la segunda parte
  lista_valores2_5 = []

  for valor_col_0, valor_col_4 in zip(tabla65.iloc[:, 0], tabla65.iloc[:, 4]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores2_5.append(valor_col_4)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores2_5)):
      if pd.isna(lista_valores2_5[i]):  # Verificar si es NaN
          lista_valores2_5[i] = "Verde máximo mínimo por detector"
      elif lista_valores2_5[i] == "1":
          lista_valores2_5[i] = "Verde máximo mínimo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores2_5[i]  # Obtener el valor de la lista en la posición i
      df65.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la tercera parte
  lista_valores3_5 = []

  for valor_col_0, valor_col_5 in zip(tabla65.iloc[:, 0], tabla65.iloc[:, 5]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores3_5.append(valor_col_5)  # Agregamos a la lista los valores de la columna 4

  for i in range(len(lista_valores3_5)):
      if pd.isna(lista_valores3_5[i]):  # Verificar si es NaN
          lista_valores3_5[i] = "Verde máximo por detector"
      elif lista_valores3_5[i] == "1":
          lista_valores3_5[i] = "Verde máximo permanente"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores3_5[i]  # Obtener el valor de la lista en la posición i
      df65.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la cuarta parte
  lista_valores4_5 = []

  for valor_col_0, valor_col_7 in zip(tabla65.iloc[:, 0], tabla65.iloc[:, 7]):
      if valor_col_0 == "NeXt":  # Detenemos el bucle cuando en la columna 0 aparece "next"
          break
      lista_valores4_5.append(valor_col_7)  # Agregamos a la lista los valores de la columna 4

  # Recorremos la lista y alteramos solo los campos de texto
  for i in range(len(lista_valores4_5)):
      if isinstance(lista_valores4_5[i], str):  # Verificamos si es una cadena de texto
          lista_valores4_5[i] = lista_valores4_5[i].split('-')[0]

  for i in range(len(lista_valores4_5)):
      if pd.isna(lista_valores4_5[i]):  # Verificar si es NaN
          lista_valores4_5[i] = "Cuenta regresiva comienza cuando un grupo en conflicto es demandado"
      elif lista_valores4_5[i] == "1":
          lista_valores4_5[i] = "Cuenta regresiva comienza sin que un grupo en conflicto sea demandado"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores4_5[i]  # Obtener el valor de la lista en la posición i
      df64.loc[fila_a_modificar, columna] = 'ok'

  ########################### Completamos la quinta parte
  lista_valores5_5 = []

  for valor_col_0, valor_col_2 in zip(tabla65.iloc[:, 0], tabla65.iloc[:, 2]):
      if valor_col_0 == "NeXt":
          break
      lista_valores5_5.append(valor_col_2)

  for i in range(len(lista_valores5_5)):
      if pd.isna(lista_valores5_5[i]):  # Verificar si es NaN
          lista_valores5_5[i] = "Luego de la extensión máxima de verde y verde pasivo"
      elif lista_valores5_5[i] == "3":
          lista_valores5_5[i] = "Luego de la extensión máxima de verde"

  for i, columna in enumerate(df6.columns):  # Empieza desde la segunda columna
      fila_a_modificar = lista_valores5_5[i]  # Obtener el valor de la lista en la posición i
      df65.loc[fila_a_modificar, columna] = 'ok'

  df65 = df65.fillna("-")



################################### Completar datos
sheet_name = "Tabla 1-6"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

valor_w1 = ws["w1"].value
estilo_w1 = ws["w1"]._style 

# Variables para la configuración de las tablas
col_start = 3  # Columna "C" (para tablas verticales)
base_row_start = 3  # Fila inicial para la primera tabla vertical
incremento_fila = 14  # Espacio entre tablas verticales

# Nuevo incremento para la fila horizontal
base_horizontal_row_start = 2  # Fila inicial para la primera fila horizontal (2)
incremento_horizontal = 14  # Incremento para la fila horizontal

# Asumiendo que `secuencias` contiene el número de tablas a llenar
for secuencia in range(1, secuencias + 1):
    # Calcular fila de inicio para la tabla horizontal y vertical
    row_start = base_row_start + (secuencia - 1) * incremento_fila
    horizontal_row_start = base_horizontal_row_start + (secuencia - 1) * incremento_horizontal

    # Seleccionar el DataFrame correspondiente para cada secuencia
    dfX = globals().get(f"df6{secuencia}", None)
    if dfX is None:
        print(f"No se encontró el DataFrame df6{secuencia}")
        continue  # Saltar a la siguiente secuencia si el DataFrame no existe

    # Rellenar la fila horizontal para el grupo actual
    for i in range(len(df2)):
        col_letra = get_column_letter(3 + i)
        ws[f"{col_letra}{horizontal_row_start}"] = df2.iloc[i, 1]

    # Rellenar la tabla verticalmente en la posición adecuada para cada secuencia
    for col_idx in range(len(dfX.columns)):
        col_letter = get_column_letter(col_start + col_idx)  # Comienza en "C" y avanza
        for row_idx in range(len(dfX)):
            if dfX.iloc[row_idx, col_idx] == "ok":
                destino = f"{col_letter}{row_start + row_idx}"
                ws[destino].value = valor_w1
                ws[destino]._style = estilo_w1 


wb.save(excel_file_path)
############################################################################################
########### Tabla 7

work_003_index = df[df.iloc[:, 0].str.contains('Work.003', na=False)].index
work_004_index = df[df.iloc[:, 0].str.contains('Work.004', na=False)].index


if not work_003_index.empty and not work_004_index.empty:
    tabla7 = df.iloc[work_003_index[0] +1 : work_004_index[0]-1] #quito el work y el Next
else:
    print("No se encontraron ambas entradas de 'Work.003' y 'Work.004'")

Tiempo_de_ciclo = tabla7['9'].tolist()

index_vals = ["Tiempo de ciclo", "Offset", "Escenario lógico", "Secuencia lógica"]

df7 = pd.DataFrame(index=index_vals, columns=range(len(Tiempo_de_ciclo)))

df7.loc["Tiempo de ciclo"] = Tiempo_de_ciclo

Offset = tabla7['10'].tolist()
df7.loc["Offset"] = Offset

work_007_index = df[df.iloc[:, 0].str.contains('Work.007', na=False)].index
work_008_index = df[df.iloc[:, 0].str.contains('Work.008', na=False)].index

if not work_007_index.empty and not work_008_index.empty:
    tabla71 = df.iloc[work_007_index[0] +1 : work_008_index[0]-1] #quito el work y el Next
else:
    print("No se encontraron ambas entradas de 'Work.997' y 'Work.998'")

escenario_logico = []
secuencia_logica = []

for i in range(len(tabla71)):
    if tabla71.iloc[i, 0] == "NeXt":  # Verificamos si la primera columna contiene "NeXt"
        if i + 1 < len(tabla71):  # Aseguramos que existe una fila después de "NeXt"
            valor_col_7 = tabla71.iloc[i + 1, 7]  # Avanzamos una fila y tomamos el valor de la columna 8
            escenario_logico.append(valor_col_7)


for i in range(len(tabla71)):
    if tabla71.iloc[i, 0] == "NeXt":
        if i + 1 < len(tabla71):
            valor_col_6 = tabla71.iloc[i + 1, 6]
            secuencia_logica.append(valor_col_6)


valor_primer_fila = df.iloc[0, 7]
escenario_logico.insert(0, valor_primer_fila)

valor_primer_fila = df.iloc[0, 6]
secuencia_logica.insert(0, valor_primer_fila)

df7.loc["Escenario lógico"] = escenario_logico
df7.loc["Secuencia lógica"] = secuencia_logica
df7 = df7.fillna("-")

################################### Completar datos
sheet_name = "Tabla 1-7"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

def escribir_fila_en_fila_excel(df, indice, fila_excel, columna_inicial):
    """
    Escribe todos los valores de una fila del DataFrame en una fila específica de Excel, comenzando en una columna dada.
    Avanza horizontalmente por las columnas.
    
    :param df: DataFrame de Pandas.
    :param indice: Índice de la fila en el DataFrame a escribir en Excel.
    :param fila_excel: Número de fila en Excel donde escribir.
    :param columna_inicial: Columna inicial en Excel (ej. "B") donde comenzar a escribir.
    """
    # Obtener los valores de la fila correspondiente al índice
    valores_fila = df.loc[indice].values  # Extrae todos los valores de la fila
    
    # Convertir la columna inicial a índice numérico
    col_inicial_idx = ord(columna_inicial.upper()) - ord('A') + 1
    
    # Escribir los valores en la fila fija en Excel, avanzando por columnas
    for i, valor in enumerate(valores_fila):
        col_letter = get_column_letter(col_inicial_idx + i)  # Calcula la letra de la columna actual
        ws[f"{col_letter}{fila_excel}"] = valor

# Escribir los valores de cada índice en las celdas correspondientes
escribir_fila_en_fila_excel(df7, "Tiempo de ciclo", 3, "B")
escribir_fila_en_fila_excel(df7, "Offset", 4, "B")
escribir_fila_en_fila_excel(df7, "Escenario lógico", 5, "B")
escribir_fila_en_fila_excel(df7, "Secuencia lógica", 6, "B")

# Guardar los cambios
wb.save(excel_file_path)

############################################################################################
########### Tabla 8

work_007_index = df[df.iloc[:, 0].str.contains('Work.007', na=False)].index
work_008_index = df[df.iloc[:, 0].str.contains('Work.008', na=False)].index

if not work_007_index.empty and not work_008_index.empty:
    tabla8 = df.iloc[work_007_index[0] +1 : work_008_index[0]] #quito el work
else:
    print("No se encontraron ambas entradas de 'Work.997' y 'Work.998'")

Num_paneles =  tabla1.iloc[4]
df8 = pd.DataFrame(index=grupo_vals, columns=range(int(Num_paneles)))

start_idx = 0  # Índice de inicio
col_num = 0    # Número de columna en df8

# Recorremos la columna 1 buscando 'NeXt'
for i, valor in enumerate(tabla8.iloc[:, 0]):
    if valor == "NeXt":  # Si encontramos 'NeXt'
        # Tomar los valores de la segunda columna entre el índice de inicio y el índice actual
        data_segment = tabla8.iloc[start_idx:i, 1].tolist()  # Extraer la segunda columna y eliminar NaN

        # Agregar esos valores como una nueva columna en df8
        df8[col_num] = data_segment

        # Actualizar el índice de inicio para la próxima búsqueda
        start_idx = i + 1

        # Incrementar el número de columna en df8
        col_num += 1

###### Vm
Num_paneles =  tabla1.iloc[4]
df81 = pd.DataFrame(index=grupo_vals, columns=range(int(Num_paneles)))

start_idx = 0  # Índice de inicio
col_num = 0    # Número de columna en df8

# Recorremos la columna 1 buscando 'NeXt'
for i, valor in enumerate(tabla8.iloc[:, 0]):
    if valor == "NeXt":  # Si encontramos 'NeXt'
        # Tomar los valores de la segunda columna entre el índice de inicio y el índice actual
        data_segment = tabla8.iloc[start_idx:i, 0].tolist()  # Extraer la segunda columna y eliminar NaN

        # Agregar esos valores como una nueva columna en df8
        df81[col_num] = data_segment

        # Actualizar el índice de inicio para la próxima búsqueda
        start_idx = i + 1

        # Incrementar el número de columna en df8
        col_num += 1

df8 = df8.fillna("-")
df81 = df81.fillna("-")

################################### Completar datos
sheet_name = "Tabla 1-8"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

# GRUPO VERTICAL
for i in range(len(df2)):
    ws[f"A{i + 4}"] = df2.iloc[i,1]

# Definir la fila de inicio y el índice de columna base (C)
fila_inicio = 4
columna_base_idx = 3  # Columna "C" en índice numérico (C = 3, E = 5, etc.)
columna_base_idx2 = 2

# Iterar sobre cada columna en df8 y colocarla en la hoja de Excel
for col_idx in range(len(df8.columns)):
    # Calcular la columna de Excel (C, E, G, etc.)
    columna_excel = get_column_letter(columna_base_idx + col_idx * 2)
    
    # Escribir los valores de la columna de df8 en Excel, comenzando desde fila_inicio y avanzando hacia abajo
    for row_idx, valor in enumerate(df8.iloc[:, col_idx]):
        ws[f"{columna_excel}{fila_inicio + row_idx}"] = valor

# Iterar sobre cada columna en df8 y colocarla en la hoja de Excel
for col_idx in range(len(df81.columns)):
    # Calcular la columna de Excel (C, E, G, etc.)
    columna_excel = get_column_letter(columna_base_idx2 + col_idx * 2)
    
    # Escribir los valores de la columna de df8 en Excel, comenzando desde fila_inicio y avanzando hacia abajo
    for row_idx, valor in enumerate(df81.iloc[:, col_idx]):
        ws[f"{columna_excel}{fila_inicio + row_idx}"] = valor

# Guardar los cambios
wb.save(excel_file_path)

############################################################################################
########### Tabla 9

work_022_index = df[df.iloc[:, 0].str.contains('Work.022', na=False)].index
work_023_index = df[df.iloc[:, 0].str.contains('Work.023', na=False)].index

if not work_022_index.empty and not work_023_index.empty:
    tabla9 = df.iloc[work_022_index[0] +1 : work_023_index[0]-1] #quito el work
else:
    print("No se encontraron ambas entradas de 'Work.022' y 'Work.998'")

##### Se sumo esta solucion para poder trabajar con la version 1 y 2 de ITC 
check_keywords = tabla9.iloc[0:,0].tolist()

keywords = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday', 'week', 'weekend']

# Función para verificar si hay coincidencias en una lista
def check_and_drop_first_column(lst, df):
    # Convertimos todos los elementos a minúsculas para una comparación insensible a mayúsculas/minúsculas
    lst = [str(item).lower() for item in lst]
    
    # Comprobar si alguno de los keywords está en la lista
    if any(keyword in item for item in lst for keyword in keywords):
        # Si hay coincidencias, eliminar la primera columna del DataFrame
        df.drop(df.columns[0], axis=1, inplace=True)
        return True
    return False

# Llamar a la función
if check_and_drop_first_column(check_keywords, tabla9):
    print("Se eliminó la primera columna de la tabla 9.")
else:
    print("No se encontró ninguna coincidencia.")
#####

lunes = tabla9.iloc[0].tolist()
lunes = [x for x in lunes if pd.notna(x)]

M_J = tabla9.iloc[1].tolist()
M_J = [x for x in M_J if pd.notna(x)]

viernes = tabla9.iloc[2].tolist()
viernes = [x for x in viernes if pd.notna(x)]

sabado = tabla9.iloc[3].tolist()
sabado = [x for x in sabado if pd.notna(x)]

domingo = tabla9.iloc[4].tolist()
domingo = [x for x in domingo if pd.notna(x)]

# Definir las columnas y el índice
columnas = ["lunes", "martes a jueves", "viernes", "sábado", "domingo"]

# Generar los índices de tiempo de 30 minutos desde 00:00 hasta 23:30
indice = pd.date_range("00:00", "23:30", freq="15min").strftime("%H:%M")

# Crear el DataFrame vacío
df9 = pd.DataFrame(index=indice, columns=columnas)

for item in lunes:
    plan, hora_inicio, minutos = item.split('-')
    hora_formato = f"{hora_inicio.zfill(2)}:{minutos.zfill(2)}"

    if hora_formato in df9.index:
        df9.at[hora_formato, 'lunes'] = f"Plan {int(plan)}"

for item in M_J:
    plan, hora_inicio, minutos = item.split('-')
    hora_formato = f"{hora_inicio.zfill(2)}:{minutos.zfill(2)}"

    if hora_formato in df9.index:
        df9.at[hora_formato, 'martes a jueves'] = f"Plan {int(plan)}"

for item in viernes:
    plan, hora_inicio, minutos = item.split('-')
    hora_formato = f"{hora_inicio.zfill(2)}:{minutos.zfill(2)}"

    if hora_formato in df9.index:
        df9.at[hora_formato, 'viernes'] = f"Plan {int(plan)}"

for item in sabado:
    plan, hora_inicio, minutos = item.split('-')
    hora_formato = f"{hora_inicio.zfill(2)}:{minutos.zfill(2)}"

    if hora_formato in df9.index:
        df9.at[hora_formato, 'sábado'] = f"Plan {int(plan)}"

for item in domingo:
    plan, hora_inicio, minutos = item.split('-')
    hora_formato = f"{hora_inicio.zfill(2)}:{minutos.zfill(2)}"

    if hora_formato in df9.index:
        df9.at[hora_formato, 'domingo'] = f"Plan {int(plan)}"

################################### Completar datos
sheet_name = "Tabla 1-9"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]


# Aseguramos que df9.index se trate como lista y eliminamos ceros iniciales en horas
index_list = [time.lstrip("0") if time.startswith("0") else time for time in df9.index]

# Iterar sobre las columnas del DataFrame
for col_idx in range(df9.shape[1]):
    column_values = df9.iloc[:, col_idx]
    
    for row_idx, row_value in enumerate(column_values):  # Convertimos a enumerate para obtener el índice
        if pd.notna(row_value):  # Verificar si el valor no es NaN
            # Convertimos el índice en cadena sin ceros iniciales para comparar con el Excel
            indice = str(index_list[row_idx])
            valor = row_value

            # Buscar coincidencia en la columna A del Excel
            for row in range(1, ws.max_row + 1):
                cell_value = str(ws[f"A{row}"].value).lstrip("0")  # También eliminamos ceros iniciales del valor en Excel
                
                if cell_value == indice:  # Coincidencia encontrada en columna A
                    # Escribir en la columna correspondiente (B, C, D, etc.) según el índice de columna en df
                    ws[f"{chr(66 + col_idx)}{row}"] = valor
                    break  # Detener el bucle interno al encontrar la coincidencia

# Guardar los cambios en el archivo Excel
wb.save(excel_file_path)


############################################################################################
########### Tabla 10

work_012_index = df[df.iloc[:, 0].str.contains('Work.012', na=False)].index
work_013_index = df[df.iloc[:, 0].str.contains('Work.013', na=False)].index

if not work_012_index.empty and not work_013_index.empty:
    tabla10 = df.iloc[work_012_index[0] +1 : work_013_index[0]-1] #quito el work
else:
    print("No se encontraron ambas entradas de 'Work.997' y 'Work.998'")

cols = ["Nombre detector","Número del grupo asociado", "Nombre del grupo asociado","Extensión de verde (s)"]
detector = tabla10.iloc[:, 0].tolist()
grupo = tabla10.iloc[:, 1].tolist()
verde = tabla10.iloc[:, 13].tolist()


df10 = pd.DataFrame()

num_filas = len(detector)
df10 = pd.DataFrame(index=range(num_filas))


for cols in cols:
    df10[cols] = pd.Series([None] * num_filas)

df10["Nombre detector"] = detector
df10["Número del grupo asociado"] = grupo
df10["Extensión de verde (s)"] = verde

grupo_vals = grupo_vals.tolist()
verde = tabla10.iloc[:, 13].tolist()

for i in range(len(df10)):
    numero = df10.iloc[i, 1]  # Obtener el valor de la segunda columna (índice 1)
    if pd.notna(numero) and int(numero) < len(grupo_vals):  # Verificar que el número sea válido y esté dentro de rango
        df10.iloc[i, 2] = grupo_vals[int(numero)-1]

df10 = df10.fillna("-")     

################################### Completar datos
sheet_name = "Tabla 1-10"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

start_row = 3  # Fila inicial
start_col = 2  # Columna inicial 'B' corresponde al índice 2

for col_idx in range(min(df10.shape[1], 4)):  # Asegura que no intente escribir más de 4 columnas
    excel_col = chr(start_col + col_idx + 64)  # Convertir índice a letra ASCII, donde 65 es 'A'
    for row_idx, value in enumerate(df10.iloc[:, col_idx]):
        ws[f"{excel_col}{start_row + row_idx}"] = value

# Guardar los cambios en el archivo Excel
wb.save(excel_file_path)

#########################################################################################################################################
#PopUp final

def mostrar_popup_finalizado():
    messagebox.showinfo("Proceso finalizado", "El proceso ha sido completado exitosamente.")

# Ejemplo de ejecución al final del proceso
mostrar_popup_finalizado()